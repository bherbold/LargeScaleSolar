from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def saveToExcel (file, data):

#data needs to be three dimensions

    wb = Workbook()

    for sheet in range(len(data)):
        ws = wb.create_sheet(str(sheet))
        for row in range(len(data[sheet])+1):
            r = []
            if row == 0:
                r = r
            else:

                for value in data[sheet][row-1]:
                    if type(value) == float:
                        r.append(value)

                    else:
                        r.append(0)
            ws.append(r)

    wb.save(filename = file)

def saveToExcelInDepth (file, data):

#data needs to be three dimensions

    wb = Workbook()
   # ws = wb.active

    for sheet in range(len(data)):

        ws = wb.create_sheet(str(sheet))
        for row in range(0,8762):
            r = []
            hourCSP = []
            hourPV = []
            if row == 0:
                r = r
                ws.append(r)
            elif row == 1:

                for value in data[sheet]:
                    if type(value) == float:
                        r.append(value)
                    else:

                        r.append(0)
                ws.append(r)
            elif row > 1:
                for value in data[sheet]:
                    if type(value) == float or type(value) == int:
                        r.append(0)
                    else:
                        if len(value) == 0:
                            r.append(0)
                        else:
                           # print(type(value))
                            r.append(float(value[row-2]))


                ws.append(r)

    wb.save(filename = file)