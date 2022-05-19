from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def readExcelSimple (file, numberRows):
    wb = load_workbook(file, data_only=True)

    #number of Rows: including heading line
    # worksheets
    # ws = wb['Hourly Data']
    # print(ws['A2'].value)
    # ws['A1'].value = '000'
    # wb.save('Sheet1')
    # Excel needs to be closed and will overwrite the value in the file

    allSheets = [] #all sheets

    for plant in wb:
        ws = wb[plant.title]

        csp_cap = 0  # KW
        sf_area = 0  # m2
        thermal_storage = 0  # MWh_th
        thermal_storage_hours = 0 #hours
        receiver_power = 0  # MW_th
        tower_height = 0  # m
        el_heater = 0  # KW_th
        csp_land = 0  # m2
        pv_cap = 0  # Wdc
        batt_pp = 0  # MW
        batt_cap = 0  # MWh_e
        batt_an_gen = 0  # GWh/y
        pv_land = 0  # m2
        csp_an_gen = 0  # GWh/y (net)
        distance_grid = 0  # km
        distance_road = 0  # km
        distance_gas = 0  # km
        distance_water = 0  # km
        other_comp = 0  # unit
        pv_el_to_grid = 0  # kwh
        csp_net_gen_hourly = []  # hourly
        pv_el_hourly = []  # hourly
        solarMultiple = 0

        setup = []

        for row in range(2, numberRows+1):

            for col in range(1, 24):
                char = get_column_letter(col)
                if col == 1:
                    csp_cap = float(ws[char + str(row)].value)
                elif col == 2:
                    sf_area = float(ws[char + str(row)].value)
                elif col == 3:
                     thermal_storage = float(ws[char + str(row)].value)
                elif col == 4:
                    thermal_storage_hours = float(ws[char + str(row)].value)
                elif col == 5:
                     receiver_power= float(ws[char + str(row)].value)
                elif col == 6:
                     tower_height= float(ws[char + str(row)].value)
                elif col == 7:
                     el_heater = float(ws[char + str(row)].value)
                elif col == 8:
                     csp_land = float(ws[char + str(row)].value)
                elif col == 9:
                     pv_cap = float(ws[char + str(row)].value)
                elif col == 10:
                     batt_pp= float(ws[char + str(row)].value)
                elif col == 11:
                     batt_cap= float(ws[char + str(row)].value)
                elif col == 12:
                     batt_an_gen= float(ws[char + str(row)].value)
                elif col == 13:
                     pv_land = float(ws[char + str(row)].value)
                elif col == 14:
                     csp_an_gen= float(ws[char + str(row)].value)
                elif col == 15:
                     distance_grid= float(ws[char + str(row)].value)
                elif col == 16:
                     distance_road= float(ws[char + str(row)].value)
                elif col == 17:
                     distance_gas= float(ws[char + str(row)].value)
                elif col == 18:
                     distance_water= float(ws[char + str(row)].value)
                elif col == 19:
                     other_comp= float(ws[char + str(row)].value)
                elif col == 20:
                     pv_el_to_grid = float(ws[char + str(row)].value)
                elif col == 21:
                     csp_net_gen_hourly = float(ws[char + str(row)].value)
                elif col == 22:
                     pv_el_hourly = float(ws[char + str(row)].value)
                elif col == 23:
                     solarMultiple= float(ws[char + str(row)].value)
            setup.append([csp_cap,sf_area,thermal_storage,thermal_storage_hours,receiver_power,tower_height,el_heater,csp_land,pv_cap,
            batt_pp,batt_cap,batt_an_gen,pv_land,csp_an_gen,distance_grid,distance_road,distance_gas,
            distance_water,other_comp,pv_el_to_grid,csp_net_gen_hourly,pv_el_hourly,solarMultiple])
        allSheets.append(setup)
    return allSheets

def readExcelIndepth (file, numberCol):
    wb = load_workbook(file, data_only=True)

    #number of Rows: including heading line
    # worksheets
    # ws = wb['Hourly Data']
    # print(ws['A2'].value)
    # ws['A1'].value = '000'
    # wb.save('Sheet1')
    # Excel needs to be closed and will overwrite the value in the file

    allSheets = [] #all sheets

    for plant in wb:
        ws = wb[plant.title]

        csp_cap = 0  # KW
        sf_area = 0  # m2
        thermal_storage = 0  # MWh_th
        thermal_storage_hours = 0 #hours
        receiver_power = 0  # MW_th
        tower_height = 0  # m
        el_heater = 0  # KW_th
        csp_land = 0  # m2
        pv_cap = 0  # Wdc
        batt_pp = 0  # MW
        batt_cap = 0  # MWh_e
        batt_an_gen = 0  # GWh/y
        pv_land = 0  # m2
        csp_an_gen = 0  # GWh/y (net)
        distance_grid = 0  # km
        distance_road = 0  # km
        distance_gas = 0  # km
        distance_water = 0  # km
        other_comp = 0  # unit
        pv_el_to_grid = 0  # kwh
        csp_net_gen_hourly = []  # hourly
        pv_el_hourly = []  # hourly
        solarMultiple = 0
        TES_SOC = [] #state of charge at each hour
        TES_charge = []  # charging
        TES_dispatch = [] #dispatch of TES
        inverterCap = 0
        inverterEff = 0
        PVselfCon = 0
        TES_loss = 0
        Inverter_cap = 0


        for col in range(1, numberCol+1):
            char = get_column_letter(col)
            if col == 1:
                csp_cap = float(ws[char + str('2')].value)
            elif col == 2:
                sf_area = float(ws[char + str('2')].value)
            elif col == 3:
                 thermal_storage = float(ws[char + str('2')].value)
            elif col == 4:
                thermal_storage_hours = float(ws[char + str('2')].value)
            elif col == 5:
                 receiver_power= float(ws[char + str('2')].value)
            elif col == 6:
                 tower_height= float(ws[char + str('2')].value)
            elif col == 7:
                 el_heater = float(ws[char + str('2')].value)
            elif col == 8:
                 csp_land = float(ws[char + str('2')].value)
            elif col == 9:
                 pv_cap = float(ws[char + str('2')].value)
            elif col == 10:
                 batt_pp= float(ws[char + str('2')].value)
            elif col == 11:
                 batt_cap= float(ws[char + str('2')].value)
            elif col == 12:
                 batt_an_gen= float(ws[char + str('2')].value)
            elif col == 13:
                 pv_land = float(ws[char + str('2')].value)
            elif col == 14:
                 csp_an_gen= float(ws[char + str('2')].value)
            elif col == 15:
                 distance_grid= float(ws[char + str('2')].value)
            elif col == 16:
                 distance_road= float(ws[char + str('2')].value)
            elif col == 17:
                 distance_gas= float(ws[char + str('2')].value)
            elif col == 18:
                 distance_water= float(ws[char + str('2')].value)
            elif col == 19:
                 other_comp= float(ws[char + str('2')].value)
            elif col == 20:
                pv_el_to_grid = float(ws[char + str('2')].value)
            elif col == 21:
                for row in range(2, 8762):
                    value = float(ws[char + str(row)].value)
                    csp_net_gen_hourly.append(value* 0.42* 0.89*1000)   # x thermal eff x receiver eff x 1000 KW= kWe
            elif col == 22:
                for row in range(2, 8762):
                    value = float(ws[char + str(row)].value)
                    pv_el_hourly.append(value)  # maybe just -1 ?
            elif col == 23:
                 solarMultiple= float(ws[char + str('2')].value)
            elif col == 24:
                 inverterCap = float(ws[char + str('2')].value)
            elif col == 25:
                 inverterEff = float(ws[char + str('2')].value)
            elif col == 26:
                 PVselfCon = float(ws[char + str('2')].value)
            elif col == 27:
                TES_loss = float(ws[char + str('2')].value)
            elif col == 28:
                Inverter_cap = float(ws[char + str('2')].value)
        allSheets.append([csp_cap,sf_area,thermal_storage,thermal_storage_hours,receiver_power,tower_height,el_heater,csp_land,pv_cap,
            batt_pp,batt_cap,batt_an_gen,pv_land,csp_an_gen,distance_grid,distance_road,distance_gas,
            distance_water,other_comp,pv_el_to_grid,csp_net_gen_hourly,pv_el_hourly,solarMultiple, TES_SOC,TES_dispatch,TES_charge
                          ,inverterCap, inverterEff,PVselfCon,TES_loss,Inverter_cap])
    return allSheets